#!/usr/bin/env python3
"""Inflection Ventures Financial Model - Excel Builder"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, AreaChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.workbook.defined_name import DefinedName

# ── Colours ────────────────────────────────────────────────────────────
C_DARK_BLUE  = "0A2463"
C_MED_BLUE   = "1565C0"
C_GREEN      = "3BBA9C"
C_RED        = "E53935"
C_GRAY       = "757575"
C_LT_GRAY    = "F5F5F5"
C_YELLOW     = "FFF9E6"
C_LT_GREEN   = "E8F5E9"
C_LT_RED     = "FFEBEE"
C_WHITE      = "FFFFFF"

# ── Style helpers ───────────────────────────────────────────────────────
def fill(hex_col):
    return PatternFill(start_color=hex_col, end_color=hex_col, fill_type="solid")

def font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border(t="thin", b="thin", l="thin", r="thin"):
    def s(x): return Side(style=x) if x else Side(style=None)
    return Border(top=s(t), bottom=s(b), left=s(l), right=s(r))

def hdr_fill(): return fill(C_DARK_BLUE)
def inp_fill(): return fill(C_YELLOW)
def calc_fill(): return fill(C_LT_GREEN)
def alt_fill(i): return fill(C_WHITE) if i % 2 == 0 else fill(C_LT_GRAY)

def hdr_font(): return font(bold=True, size=11, color=C_WHITE)
def inp_font(): return font(bold=True, size=10)
def sub_hdr_font(): return font(bold=True, size=10, color=C_WHITE)
def section_font(): return font(bold=True, size=12)
def total_font(): return font(bold=True, size=10)

def style(cell, fll=None, fnt=None, aln=None, brd=None, fmt=None):
    if fll: cell.fill = fll
    if fnt: cell.font = fnt
    if aln: cell.alignment = aln
    if brd: cell.border = brd
    if fmt: cell.number_format = fmt

def banner(ws, row, col, text, span=1, size=12):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = hdr_fill(); c.font = font(bold=True, size=size, color=C_WHITE)
    c.alignment = align("left","center")
    c.border = border("medium","medium","medium","medium")
    if span>1: ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    return c

def section(ws, row, col, text, span=1, clr=None):
    clr = clr or C_MED_BLUE
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(clr); c.font = font(bold=True, size=11, color=C_WHITE)
    c.alignment = align("left","center")
    c.border = border()
    if span>1: ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    return c

def th(ws, row, col, text, span=1):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill("D9E1F2"); c.font = font(bold=True, size=10, color=C_DARK_BLUE)
    c.alignment = align("center","center"); c.border = border()
    if span>1: ws.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+span-1)
    return c

def inp(ws, row, col, value, fmt="€#,##0"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = inp_fill(); c.font = inp_font()
    c.alignment = align("right","center"); c.border = border(); c.number_format = fmt
    return c

def calc(ws, row, col, formula, fmt="€#,##0", bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = calc_fill(); c.font = font(bold=bold, size=10)
    c.alignment = align("right","center"); c.border = border(); c.number_format = fmt
    return c

def lbl(ws, row, col, text, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=("  "*indent)+text)
    c.font = font(bold=bold, size=10); c.alignment = align("left","center"); c.border = border()
    return c

def row_data(ws, row, col_start, values, fmts=None, bold=False, alt_i=None):
    for j, v in enumerate(values):
        c = ws.cell(row=row, column=col_start+j, value=v)
        if alt_i is not None: c.fill = alt_fill(alt_i)
        c.font = font(bold=bold, size=10); c.alignment = align("right","center"); c.border = border()
        if fmts and j < len(fmts): c.number_format = fmts[j]

FMT_EUR   = "€#,##0"
FMT_EUR2  = "€#,##0.00"
FMT_PCT   = "0.0%"
FMT_MULT  = '0.00"x"'
FMT_NUM   = "#,##0"
FMT_TXT   = "@"

# ═══════════════════════════════════════════════════════════════════════
# SHEET 7 – ASSUMPTIONS & INPUTS (built first so other sheets reference it)
# ═══════════════════════════════════════════════════════════════════════
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions & Inputs")
    ws.sheet_properties.tabColor = C_GRAY
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 35

    banner(ws, 1, 1, "INFLECTION VENTURES – ASSUMPTIONS & INPUTS", span=5, size=14)
    ws.row_dimensions[1].height = 28
    ws.cell(row=2,column=1,value="All yellow cells are user-adjustable. Other sheets reference values here.").font = font(italic=True,size=9,color=C_GRAY)
    ws.row_dimensions[3].height = 6

    # ── A: Fund Parameters ──
    section(ws, 4, 1, "A. FUND PARAMETERS", span=5)
    ws.row_dimensions[4].height = 20
    th(ws,5,1,"Parameter"); th(ws,5,2,"Value"); th(ws,5,5,"Notes")
    params = [
        ("Fund Name",                "Inflection Ventures Fund I",  "@",    False, "Legal entity name"),
        ("Legal Structure",          "Luxembourg ELTIF 2.0",        "@",    False, "EU regulatory framework"),
        ("Fund Size (Target €)",     3000000,                       FMT_EUR, True, "User adjustable – base case €3M"),
        ("Minimum Investment (€)",   10000,                         FMT_EUR, False, "ELTIF 2.0 minimum"),
        ("Target Investors",         "300-500",                     "@",    False, "Range"),
        ("Management Fee (%)",       0.015,                         FMT_PCT, True, "Annual, on committed capital"),
        ("Carried Interest (%)",     0.25,                          FMT_PCT, True, "Above hurdle rate"),
        ("Hurdle Rate (%)",          0.08,                          FMT_PCT, True, "Annual preferred return"),
        ("Fund Duration (years)",    10,                            FMT_NUM, False, "Plus possible extensions"),
        ("Investment Period (years)",4,                             FMT_NUM, False, "Active deployment window"),
    ]
    for i,(p,v,fmt,editable,note) in enumerate(params):
        r = 6+i
        ws.row_dimensions[r].height = 18
        ws.cell(row=r,column=1,value=p).font = font(size=10)
        ws.cell(row=r,column=1).border = border()
        c = ws.cell(row=r,column=2,value=v)
        c.number_format=fmt; c.alignment=align("right","center"); c.border=border()
        if editable: c.fill=inp_fill(); c.font=inp_font()
        else: c.font=font(size=10)
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY)
        ws.cell(row=r,column=5).border=border()
    ws.row_dimensions[16].height = 6

    # Named ranges (for cross-sheet use)
    # FundSize=B8, MgmtFee=B11, CarryRate=B12, HurdleRate=B13
    def named(name, ref):
        dn = DefinedName(name=name, attr_text=f"'Assumptions & Inputs'!{ref}")
        wb.defined_names[name] = dn
    named("FundSize",    "$B$8")
    named("MgmtFee",     "$B$11")
    named("CarryRate",   "$B$12")
    named("HurdleRate",  "$B$13")

    # ── B: Portfolio Parameters ──
    section(ws, 17, 1, "B. PORTFOLIO PARAMETERS", span=5)
    ws.row_dimensions[17].height = 20
    th(ws,18,1,"Parameter"); th(ws,18,2,"Value"); th(ws,18,5,"Notes")
    pparams = [
        ("Target Companies (total)",       18,   FMT_NUM, True,  "Range 15–20"),
        ("Longevity Companies",             9,   FMT_NUM, False, "~50% of portfolio"),
        ("Space Companies",                 9,   FMT_NUM, False, "~50% of portfolio"),
        ("Average Initial Ticket (€)",  200000,  FMT_EUR, False, "€150K–€250K range"),
        ("Follow-On Reserve (%)",          0.30,  FMT_PCT, True,  "30% of fund reserved"),
        ("Geographic – EU (%)",            0.70,  FMT_PCT, False, "ELTIF 2.0 compliance"),
        ("Geographic – US/Other (%)",      0.30,  FMT_PCT, False, ""),
        ("Stage Focus",       "Pre-seed, Seed, Series A", "@", False, "Early-stage only"),
    ]
    for i,(p,v,fmt,editable,note) in enumerate(pparams):
        r = 19+i
        ws.row_dimensions[r].height = 18
        ws.cell(row=r,column=1,value=p).font=font(size=10); ws.cell(row=r,column=1).border=border()
        c=ws.cell(row=r,column=2,value=v); c.number_format=fmt; c.alignment=align("right","center"); c.border=border()
        if editable: c.fill=inp_fill(); c.font=inp_font()
        else: c.font=font(size=10)
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()
    ws.row_dimensions[27].height=6

    # ── C: Return Assumptions ──
    section(ws, 28, 1, "C. RETURN ASSUMPTIONS", span=5)
    ws.row_dimensions[28].height=20
    th(ws,29,1,"Scenario"); th(ws,29,2,"Portfolio MOIC"); th(ws,29,3,"Timeline (yrs)"); th(ws,29,5,"Notes")
    ret_data = [
        ("Conservative",  2.0, "8–9", "Base case survival rate"),
        ("Base Case",     3.0,  7,    "Target returns"),
        ("Optimistic",    5.0,  6,    "Strong portfolio"),
        ("Failures (40-50%)",     "0–0.5x", "",  "Total loss or near-loss"),
        ("Modest Returns (30-40%)","0.5–2x", "",  "Return capital or small gain"),
        ("Good Returns (10-20%)","3–10x",    "",  "Solid returns"),
        ("Exceptional (5-10%)",  "10x+",     "",  "Home runs"),
    ]
    for i,(s,m,t,n) in enumerate(ret_data):
        r=30+i; ws.row_dimensions[r].height=18
        ws.cell(row=r,column=1,value=s).font=font(size=10); ws.cell(row=r,column=1).border=border()
        ws.cell(row=r,column=2,value=m).font=font(size=10); ws.cell(row=r,column=2).border=border(); ws.cell(row=r,column=2).alignment=align("right","center")
        ws.cell(row=r,column=3,value=t).font=font(size=10); ws.cell(row=r,column=3).border=border(); ws.cell(row=r,column=3).alignment=align("center","center")
        ws.cell(row=r,column=5,value=n).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()
    ws.row_dimensions[37].height=6

    # ── D: Operating Expense Assumptions ──
    section(ws, 38, 1, "D. OPERATING EXPENSE ASSUMPTIONS", span=5)
    ws.row_dimensions[38].height=20
    for col,hd in enumerate(["Category","Year 1","Year 2","Year 3","Growth Rate"],1):
        th(ws,39,col,hd)
    expenses = [
        ("Legal & Compliance",      25000, 25000, 25000, "Flat"),
        ("Fund Administration",     18000, 22000, 25000, "+15–20%"),
        ("Auditor",                 10000, 12000, 15000, "+20%"),
        ("ELTIF Depositary",         8000, 10000, 12000, "+20%"),
        ("Platform & Technology",   12000, 12000, 15000, "+0–25%"),
        ("KYC/AML",                  8000,  5000,  3000, "Decreasing"),
        ("Marketing & Events",       8000, 10000, 12000, "+20–25%"),
        ("Insurance (D&O, E&O)",     3000,  3000,  3000, "Flat"),
        ("Bank Fees",                1500,  1500,  1500, "Flat"),
        ("Other / Contingency",      5000,  5000,  5000, "Flat"),
    ]
    exp_start=40
    for i,(cat,y1,y2,y3,gr) in enumerate(expenses):
        r=exp_start+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        ws.cell(row=r,column=1,value=cat).font=font(size=10); ws.cell(row=r,column=1).fill=fl; ws.cell(row=r,column=1).border=border()
        for j,v in enumerate([y1,y2,y3],2):
            c=ws.cell(row=r,column=j,value=v); c.number_format=FMT_EUR; c.fill=fl; c.font=font(size=10); c.alignment=align("right","center"); c.border=border()
        ws.cell(row=r,column=5,value=gr).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()

    # Total row
    tot_r=exp_start+len(expenses); ws.row_dimensions[tot_r].height=18
    ws.cell(row=tot_r,column=1,value="TOTAL EXPENSES").font=total_font(); ws.cell(row=tot_r,column=1).border=border("medium","medium","medium","medium"); ws.cell(row=tot_r,column=1).fill=fill("D9E1F2")
    for j,col in enumerate(["B","C","D"],2):
        c=ws.cell(row=tot_r,column=j,value=f"=SUM({col}{exp_start}:{col}{tot_r-1})")
        c.number_format=FMT_EUR; c.font=total_font(); c.alignment=align("right","center")
        c.border=border("medium","medium","medium","medium"); c.fill=fill("D9E1F2")
    ws.row_dimensions[tot_r+1].height=6
    # tot_r should be 50 → Expenses totals at B50,C50,D50

    # ── E: Fundraising Timeline ──
    section(ws, tot_r+2, 1, "E. FUNDRAISING TIMELINE", span=5)
    ws.row_dimensions[tot_r+2].height=20
    for col,hd in enumerate(["Milestone","Date","Amount","Investors","Notes"],1):
        th(ws,tot_r+3,col,hd)
    timeline=[
        ("First Close",  "Q2 2026","€500K – €1M",   "50–100","Minimum to proceed"),
        ("Second Close", "Q3 2026","+€1,000,000",    "+100",  ""),
        ("Final Close",  "Q4 2026","€3,000,000 total","200–300","Target"),
        ("Stretch Goal", "Q1 2027","€5,000,000",      "300–500","If strong demand"),
    ]
    for i,(m,d,a,inv,n) in enumerate(timeline):
        r=tot_r+4+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        for j,v in enumerate([m,d,a,inv,n],1):
            c=ws.cell(row=r,column=j,value=v); c.fill=fl; c.font=font(size=10); c.border=border()
    ws.row_dimensions[tot_r+8].height=6

    # ── F: Market Data ──
    mkt_base=tot_r+9
    section(ws, mkt_base, 1, "F. MARKET DATA (Reference)", span=5)
    ws.row_dimensions[mkt_base].height=20
    for col,hd in enumerate(["Metric","Longevity","Space","Source/Notes"],1):
        th(ws,mkt_base+1,col,hd)
    mkt=[
        ("Market Size (2030)","€600B+","€850B+","Various analyst reports"),
        ("Annual VC Deployment","€5–6B","€6–8B","2024 estimates"),
        ("CAGR","15–20%","10–12%",""),
        ("Notable Investors","Bezos €3B, Altman €180M","SpaceX €180B valuation",""),
    ]
    for i,(m,l,s,n) in enumerate(mkt):
        r=mkt_base+2+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        for j,v in enumerate([m,l,s,n],1):
            c=ws.cell(row=r,column=j,value=v); c.fill=fl; c.font=font(size=10); c.border=border()
    ws.row_dimensions[mkt_base+6].height=6

    # ── G: Competitive Benchmarks ──
    bench_base=mkt_base+7
    section(ws, bench_base, 1, "G. COMPETITIVE BENCHMARKS", span=5)
    ws.row_dimensions[bench_base].height=20
    for col,hd in enumerate(["Fund","Min Investment","Mgmt Fee","Carry","Our Advantage"],1):
        th(ws,bench_base+1,col,hd)
    bench=[
        ("Traditional VC","€500K – €5M","2%","20%","50–100x more accessible"),
        ("Apollo Health (Longevity)","€1M+","2%","20%","Access + dual sector"),
        ("Seraphim (Space)","€500K+","2%","20%","Access + dual sector"),
        ("Crowdfunding","€100 – €1K","Varies","N/A","Portfolio + expertise"),
        ("★ Inflection Ventures","€10,000","1.5%","25%","Unique positioning"),
    ]
    for i,(f,m,mf,c,a) in enumerate(bench):
        r=bench_base+2+i; ws.row_dimensions[r].height=18
        fl=fill("E8F0FE") if i==4 else alt_fill(i)
        bld=True if i==4 else False
        for j,v in enumerate([f,m,mf,c,a],1):
            c2=ws.cell(row=r,column=j,value=v); c2.fill=fl; c2.font=font(bold=bld,size=10); c2.border=border()

    print("  ✓ Sheet 7: Assumptions & Inputs")
    return ws, tot_r   # return expense total row so Fund Economics can reference it


# ═══════════════════════════════════════════════════════════════════════
# SHEET 2 – FUND ECONOMICS
# ═══════════════════════════════════════════════════════════════════════
def build_fund_economics(wb):
    ws = wb.create_sheet("Fund Economics")
    ws.sheet_properties.tabColor = C_MED_BLUE
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 18

    banner(ws,1,1,"FUND ECONOMICS",span=5,size=14); ws.row_dimensions[1].height=28
    ws.row_dimensions[2].height=6

    # ── A: Fund Terms ──
    section(ws,3,1,"A. FUND TERMS — INPUT SECTION",span=5); ws.row_dimensions[3].height=20
    th(ws,4,1,"Parameter"); th(ws,4,2,"Value"); th(ws,4,4,"Notes")

    fund_terms = [
        ("Fund Size (€)",          "=FundSize",      FMT_EUR, "Linked from Assumptions"),
        ("Management Fee (%)",     "=MgmtFee",       FMT_PCT, "Annual on committed capital"),
        ("Carried Interest (%)",   "=CarryRate",      FMT_PCT, "Above hurdle rate"),
        ("Hurdle Rate (%)",        "=HurdleRate",     FMT_PCT, "Annual preferred return"),
        ("Fund Duration (years)",  10,                FMT_NUM, "Plus possible extensions"),
        ("Investment Period (yrs)",4,                 FMT_NUM, "Active deployment"),
    ]
    for i,(p,v,fmt,note) in enumerate(fund_terms):
        r=5+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        ws.cell(row=r,column=1,value=p).font=font(size=10); ws.cell(row=r,column=1).fill=fl; ws.cell(row=r,column=1).border=border()
        c=ws.cell(row=r,column=2,value=v); c.number_format=fmt; c.alignment=align("right","center"); c.border=border()
        c.fill=inp_fill() if i<4 else fl; c.font=inp_font() if i<4 else font(size=10)
        ws.cell(row=r,column=4,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=4).border=border()
    # B5=FundSize, B6=MgmtFee, B7=CarryRate, B8=HurdleRate

    ws.row_dimensions[11].height=6
    section(ws,12,1,"CALCULATED FUND METRICS",span=5); ws.row_dimensions[12].height=18
    calc_items=[
        ("Annual Management Fee","=B5*B6",FMT_EUR),
        ("Total Mgmt Fees (10 years)","=B13*B9",FMT_EUR),
        ("Capital Available to Invest","=B5*(1-0.02)",FMT_EUR),  # approx 2% setup costs
    ]
    for i,(lbl_txt,formula,fmt) in enumerate(calc_items):
        r=13+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,lbl_txt,bold=True)
        calc(ws,r,2,formula,fmt=fmt,bold=True)
        lbl(ws,r,4,"→ Auto-calculated from inputs above")
    ws.row_dimensions[16].height=6

    # ── B: LP Economics Calculator ──
    section(ws,17,1,"B. LP ECONOMICS CALCULATOR",span=5); ws.row_dimensions[17].height=20
    ws.cell(row=18,column=1,value="Adjust the yellow cells to model different investor scenarios").font=font(italic=True,size=9,color=C_GRAY)
    ws.row_dimensions[18].height=14; ws.row_dimensions[19].height=6

    th(ws,20,1,"Scenario Input"); th(ws,20,2,"Value"); th(ws,20,4,"What it means")
    lp_inp=[
        ("Investment Amount (€)",10000,FMT_EUR,"Minimum: €10,000"),
        ("Portfolio MOIC",3.0,FMT_MULT,"Expected gross fund multiple"),
        ("Years to Exit",7,FMT_NUM,"Average holding period"),
    ]
    for i,(p,v,fmt,note) in enumerate(lp_inp):
        r=21+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,p)
        inp(ws,r,2,v,fmt=fmt)
        ws.cell(row=r,column=4,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=4).border=border()

    # B21=Inv Amount, B22=MOIC, B23=Years
    ws.row_dimensions[24].height=10
    section(ws,25,1,"Portfolio Returns",span=5,clr=C_GRAY); ws.row_dimensions[25].height=18
    port_calcs=[
        ("Gross Return (before fees)","=B21*B22",FMT_EUR),
        ("Gross Profit","=B26-B21",FMT_EUR),
    ]
    for i,(t,f,fmt) in enumerate(port_calcs):
        r=26+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,t); calc(ws,r,2,f,fmt=fmt)

    ws.row_dimensions[28].height=10
    section(ws,29,1,"Hurdle & Carry Calculation",span=5,clr=C_GRAY); ws.row_dimensions[29].height=18
    hurdle_calcs=[
        ("Hurdle Amount (8% × years, compound)","=B21*((1+B8)^B23)-B21",FMT_EUR),
        ("Profit Above Hurdle","=MAX(0,B27-B30)",FMT_EUR),
        ("GP Carry (25% of excess)","=B31*B7",FMT_EUR),
        ("LP Share of Excess (75%)","=B31*(1-B7)",FMT_EUR),
    ]
    for i,(t,f,fmt) in enumerate(hurdle_calcs):
        r=30+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,t); calc(ws,r,2,f,fmt=fmt)
    # B30=Hurdle, B31=ProfitAboveHurdle, B32=GP Carry, B33=LP Share

    ws.row_dimensions[34].height=10
    section(ws,35,1,"LP Total Return",span=5,clr="1B5E20"); ws.row_dimensions[35].height=18
    lp_ret=[
        ("Return of Capital","=B21",FMT_EUR),
        ("Hurdle Return","=B30",FMT_EUR),
        ("75% of Excess Profit","=B33",FMT_EUR),
        ("LP TOTAL DISTRIBUTION","=B36+B37+B38",FMT_EUR),
        ("LP Net Multiple","=B39/B21",FMT_MULT),
        ("LP Net IRR (approx.)","=(B39/B21)^(1/B23)-1",FMT_PCT),
    ]
    for i,(t,f,fmt) in enumerate(lp_ret):
        r=36+i; ws.row_dimensions[r].height=18
        is_key = i>=3
        lbl(ws,r,1,t,bold=is_key)
        c=ws.cell(row=r,column=2,value=f); c.number_format=fmt
        c.alignment=align("right","center"); c.border=border()
        if i==3:
            c.fill=fill("C8E6C9"); c.font=font(bold=True,size=14,color="1B5E20")
        elif i in [4,5]:
            c.fill=fill("C8E6C9"); c.font=font(bold=True,size=12,color="1B5E20")
        else:
            c.fill=calc_fill(); c.font=font(size=10)
    ws.row_dimensions[41].height=6

    # ── C: GP Economics ──
    section(ws,42,1,"C. GP ECONOMICS",span=5); ws.row_dimensions[42].height=20
    th(ws,43,1,"Revenue Stream"); th(ws,43,2,"Amount"); th(ws,43,4,"Basis")

    gp_items=[
        ("Annual Management Fee","=B13",FMT_EUR,"Fund Size × 1.5%"),
        ("Total Mgmt Fees (10 years)","=B14",FMT_EUR,"Annual Fee × 10"),
        ("","","",""),
        ("Fund Size (for carry calc)","=B5",FMT_EUR,"From inputs"),
        ("Fund MOIC assumed","=B22",FMT_MULT,"From LP scenario above"),
        ("Gross Fund Returns","=B5*B22",FMT_EUR,"Fund Size × MOIC"),
        ("Gross Fund Profit","=B49-B5",FMT_EUR,"Returns − Fund Size"),
        ("Hurdle to LPs (8%, 7yr avg)","=B5*((1+B8)^7)-B5",FMT_EUR,"Preferred return on entire fund"),
        ("Profit Above Hurdle","=MAX(0,B50-B51)",FMT_EUR,""),
        ("GP Carry (25%)","=B52*B7",FMT_EUR,"Carry on profit above hurdle"),
        ("","","",""),
        ("TOTAL GP COMPENSATION","=B44+B53",FMT_EUR,"Mgmt Fees + Carry"),
    ]
    for i,(t,f,fmt,note) in enumerate(gp_items):
        r=44+i; ws.row_dimensions[r].height=18
        if t:
            lbl(ws,r,1,t,bold=(i in [0,1,11]))
            c=ws.cell(row=r,column=2,value=f); c.number_format=fmt
            c.alignment=align("right","center"); c.border=border()
            if i==11:
                c.fill=fill("C8E6C9"); c.font=font(bold=True,size=12,color="1B5E20")
            else:
                c.fill=calc_fill(); c.font=font(size=10)
            ws.cell(row=r,column=4,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=4).border=border()

    ws.row_dimensions[56].height=6

    # ── D: Fee Comparison Table ──
    section(ws,57,1,"D. FEE COMPARISON TABLE",span=5); ws.row_dimensions[57].height=20
    for col,hd in enumerate(["Fund Size","Mgmt Fee (1.5%)","Annual Revenue","10-Year Total"],1):
        th(ws,58,col,hd)
    fee_sizes=[(1000000,"€1M"),(2000000,"€2M"),(3000000,"€3M"),(5000000,"€5M")]
    for i,(sz,lbl_txt) in enumerate(fee_sizes):
        r=59+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        ws.cell(row=r,column=1,value=lbl_txt).font=font(size=10); ws.cell(row=r,column=1).fill=fl; ws.cell(row=r,column=1).border=border()
        ann=sz*0.015
        for j,v in enumerate([sz*0.015,ann,ann*10],2):
            c=ws.cell(row=r,column=j,value=v); c.number_format=FMT_EUR; c.fill=fl; c.font=font(size=10)
            c.alignment=align("right","center"); c.border=border()

    print("  ✓ Sheet 2: Fund Economics")
    return ws


# ═══════════════════════════════════════════════════════════════════════
# SHEET 3 – 3-YEAR OPERATING PLAN
# ═══════════════════════════════════════════════════════════════════════
def build_operating_plan(wb):
    ws = wb.create_sheet("3-Year Operating Plan")
    ws.sheet_properties.tabColor = C_GREEN
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    for col in ["B","C","D"]: ws.column_dimensions[col].width = 16
    ws.column_dimensions["E"].width = 30

    banner(ws,1,1,"3-YEAR OPERATING PLAN",span=5,size=14); ws.row_dimensions[1].height=28
    ws.row_dimensions[2].height=6

    # ── A: Revenue ──
    section(ws,3,1,"A. REVENUE",span=5); ws.row_dimensions[3].height=20
    for col,hd in enumerate(["Line Item","Year 1","Year 2","Year 3","Notes"],1):
        th(ws,4,col,hd)

    # AUM flow
    # Capital raised inputs (editable)
    aum_rows=[
        ("Beginning AUM (€)",      None,               "=D10",         "=E10",         "Carried from prior year"),
        ("(+) Capital Raised (€)",  2000000,            2000000,        1000000,        "User input – adjust as needed"),
        ("(−) Capital Deployed (€)",-500000,           -1500000,       -2000000,       "Negative = outflow to portfolio"),
        ("Ending AUM (€)",         "=D5+D6+D7",        "=E5+E6+E7",   "=F5+F6+F7",   "Beginning + Raised + Deployed"),
        ("Average AUM (€)",        "=(D5+D8)/2",       "=(E5+E8)/2",  "=(F5+F8)/2",  "(Beginning + Ending) / 2"),
    ]
    for i,(lbl_txt,y1,y2,y3,note) in enumerate(aum_rows):
        r=5+i; ws.row_dimensions[r].height=18; is_inp=(i==1 or i==2); fl=alt_fill(i)
        lbl(ws,r,1,lbl_txt,bold=(i in [0,3,4]))
        ws.cell(row=r,column=1).fill=fl
        for j,(col_l,v) in enumerate(zip(["B","C","D"],[y1,y2,y3])):
            c=ws.cell(row=r,column=2+j,value=v); c.number_format=FMT_EUR
            c.alignment=align("right","center"); c.border=border()
            if is_inp and v is not None and not isinstance(v,str):
                c.fill=inp_fill(); c.font=inp_font()
            else:
                c.fill=calc_fill() if not fl else fl; c.font=font(bold=(i in [0,3,4]),size=10)
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()

    # Fix Year1 beginning AUM = 0
    c=ws["B5"]; c.value=0; c.fill=calc_fill(); c.font=font(size=10)
    # Make capital raised/deployed editable
    for r in [6,7]:
        for col in ["B","C","D"]:
            c=ws[f"{col}{r}"]; c.fill=inp_fill(); c.font=inp_font()

    ws.row_dimensions[10].height=6
    section(ws,11,1,"MANAGEMENT FEE REVENUE",span=5,clr=C_GRAY); ws.row_dimensions[11].height=18
    mgmt_rows=[
        ("Management Fee Revenue (1.5%)", "=B9*'Fund Economics'!B6","=C9*'Fund Economics'!B6","=D9*'Fund Economics'!B6","Avg AUM × 1.5%"),
        ("Total Revenue",                 "=B12",                   "=C12",                   "=D12",                    "= Management Fee"),
    ]
    for i,(lbl_txt,y1,y2,y3,note) in enumerate(mgmt_rows):
        r=12+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,lbl_txt,bold=True)
        for j,v in enumerate([y1,y2,y3],2):
            calc(ws,r,j,v,fmt=FMT_EUR,bold=True)
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()

    ws.row_dimensions[14].height=6

    # ── B: Operating Expenses ──
    section(ws,15,1,"B. OPERATING EXPENSES",span=5); ws.row_dimensions[15].height=20
    for col,hd in enumerate(["Category","Year 1","Year 2","Year 3","Notes"],1):
        th(ws,16,col,hd)

    expenses_ref = [
        ("Legal & Compliance",     "='Assumptions & Inputs'!B40","='Assumptions & Inputs'!C40","='Assumptions & Inputs'!D40","Fund counsel, regulatory"),
        ("Fund Administration",    "='Assumptions & Inputs'!B41","='Assumptions & Inputs'!C41","='Assumptions & Inputs'!D41","NAV, reporting, accounting"),
        ("Auditor",                "='Assumptions & Inputs'!B42","='Assumptions & Inputs'!C42","='Assumptions & Inputs'!D42","Annual audit"),
        ("ELTIF Depositary",       "='Assumptions & Inputs'!B43","='Assumptions & Inputs'!C43","='Assumptions & Inputs'!D43","Custody services"),
        ("Platform & Technology",  "='Assumptions & Inputs'!B44","='Assumptions & Inputs'!C44","='Assumptions & Inputs'!D44","Investor portal, hosting"),
        ("KYC/AML",                "='Assumptions & Inputs'!B45","='Assumptions & Inputs'!C45","='Assumptions & Inputs'!D45","Onboarding (higher Y1)"),
        ("Marketing & Events",     "='Assumptions & Inputs'!B46","='Assumptions & Inputs'!C46","='Assumptions & Inputs'!D46","Investor acquisition"),
        ("Insurance (D&O, E&O)",   "='Assumptions & Inputs'!B47","='Assumptions & Inputs'!C47","='Assumptions & Inputs'!D47",""),
        ("Bank Fees",              "='Assumptions & Inputs'!B48","='Assumptions & Inputs'!C48","='Assumptions & Inputs'!D48","Account maintenance"),
        ("Other / Contingency",    "='Assumptions & Inputs'!B49","='Assumptions & Inputs'!C49","='Assumptions & Inputs'!D49","Buffer"),
    ]
    exp_start=17
    for i,(cat,y1,y2,y3,note) in enumerate(expenses_ref):
        r=exp_start+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        lbl(ws,r,1,cat); ws.cell(row=r,column=1).fill=fl
        for j,v in enumerate([y1,y2,y3],2):
            c=ws.cell(row=r,column=j,value=v); c.number_format=FMT_EUR; c.fill=fl
            c.font=font(size=10); c.alignment=align("right","center"); c.border=border()
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()

    tot_exp_r=exp_start+len(expenses_ref); ws.row_dimensions[tot_exp_r].height=18
    lbl(ws,tot_exp_r,1,"TOTAL EXPENSES",bold=True); ws.cell(row=tot_exp_r,column=1).fill=fill("D9E1F2"); ws.cell(row=tot_exp_r,column=1).border=border("medium","medium","medium","medium")
    for j,col in enumerate(["B","C","D"],2):
        c=ws.cell(row=tot_exp_r,column=j,value=f"=SUM({col}{exp_start}:{col}{tot_exp_r-1})")
        c.number_format=FMT_EUR; c.font=total_font(); c.alignment=align("right","center")
        c.border=border("medium","medium","medium","medium"); c.fill=fill("D9E1F2")
    # tot_exp_r = 27

    ws.row_dimensions[tot_exp_r+1].height=6

    # ── C: Cash Flow Summary ──
    cf_base=tot_exp_r+2
    section(ws,cf_base,1,"C. CASH FLOW SUMMARY",span=5); ws.row_dimensions[cf_base].height=20
    for col,hd in enumerate(["Line Item","Year 1","Year 2","Year 3"],1):
        th(ws,cf_base+1,col,hd)

    cf_rows=[
        ("Revenue (Mgmt Fees)",    f"=B13",f"=C13",f"=D13"),
        ("Operating Expenses",     f"=B{tot_exp_r}",f"=C{tot_exp_r}",f"=D{tot_exp_r}"),
        ("Net Operating Cash Flow",f"=B{cf_base+2}-B{cf_base+3}",f"=C{cf_base+2}-C{cf_base+3}",f"=D{cf_base+2}-D{cf_base+3}"),
        ("Cumulative Cash Flow",   f"=B{cf_base+4}",f"=B{cf_base+5}+C{cf_base+4}",f"=C{cf_base+5}+D{cf_base+4}"),
    ]
    for i,(lbl_txt,y1,y2,y3) in enumerate(cf_rows):
        r=cf_base+2+i; ws.row_dimensions[r].height=18
        is_key=(i>=2)
        lbl(ws,r,1,lbl_txt,bold=is_key)
        for j,v in enumerate([y1,y2,y3],2):
            calc(ws,r,j,v,fmt=FMT_EUR,bold=is_key)

    # Gap coverage
    ws.row_dimensions[cf_base+6].height=8
    section(ws,cf_base+7,1,"FUNDING GAP COVERAGE",span=5,clr=C_GRAY); ws.row_dimensions[cf_base+7].height=18
    gap_rows=[
        ("Grant Funding (target)",     30000, FMT_EUR, True),
        ("GP Capital Contribution",    20000, FMT_EUR, True),
        ("Total Gap Coverage",        "=B{0}+B{1}".format(cf_base+8,cf_base+9), FMT_EUR, False),
        ("Adjusted Cumulative CF (Y3)",f"=D{cf_base+5}+B{cf_base+10}", FMT_EUR, False),
    ]
    for i,(lbl_txt,v,fmt,editable) in enumerate(gap_rows):
        r=cf_base+8+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,lbl_txt,bold=(i>=2))
        c=ws.cell(row=r,column=2,value=v); c.number_format=fmt
        c.alignment=align("right","center"); c.border=border()
        if editable: c.fill=inp_fill(); c.font=inp_font()
        else: c.fill=calc_fill(); c.font=font(bold=(i>=2),size=10)

    ws.row_dimensions[cf_base+12].height=6

    # ── D: Break-Even Analysis ──
    be_base=cf_base+13
    section(ws,be_base,1,"D. BREAK-EVEN ANALYSIS",span=5); ws.row_dimensions[be_base].height=20
    ws.row_dimensions[be_base+1].height=18
    lbl(ws,be_base+1,1,"Average Annual Operating Expenses")
    calc(ws,be_base+1,2,f"=AVERAGE(B{tot_exp_r},C{tot_exp_r},D{tot_exp_r})",fmt=FMT_EUR)
    ws.row_dimensions[be_base+2].height=18
    lbl(ws,be_base+2,1,"Break-Even AUM (÷ 1.5% mgmt fee)")
    c=calc(ws,be_base+2,2,f"=B{be_base+1}/MgmtFee",fmt=FMT_EUR,bold=True)
    c.font=font(bold=True,size=14,color=C_MED_BLUE)
    ws.row_dimensions[be_base+3].height=18
    ws.cell(row=be_base+3,column=1,value='→ Fund reaches operational break-even at this AUM level').font=font(italic=True,size=10,color=C_MED_BLUE)
    ws.merge_cells(start_row=be_base+3,start_column=1,end_row=be_base+3,end_column=5)

    # ── Charts ──
    ws.row_dimensions[be_base+5].height=6

    # Hidden chart data
    chart_base=be_base+6
    ws.cell(row=chart_base,column=1,value="CHART DATA (for charts below)").font=font(bold=True,size=9,color=C_GRAY)
    for col,hd in enumerate(["","Year 1","Year 2","Year 3"],1):
        ws.cell(row=chart_base+1,column=col,value=hd).font=font(bold=True,size=9)
    ws.cell(row=chart_base+2,column=1,value="Revenue").font=font(size=9)
    ws.cell(row=chart_base+3,column=1,value="Expenses").font=font(size=9)
    ws.cell(row=chart_base+4,column=1,value="Net CF").font=font(size=9)
    ws.cell(row=chart_base+5,column=1,value="AUM").font=font(size=9)
    for j in range(3):
        col=j+2
        ws.cell(row=chart_base+2,column=col,value=f"=B13" if j==0 else f"={'C' if j==1 else 'D'}13")
        ws.cell(row=chart_base+3,column=col,value=f"=B{tot_exp_r}" if j==0 else f"={'C' if j==1 else 'D'}{tot_exp_r}")
        ws.cell(row=chart_base+4,column=col,value=f"=B{cf_base+4}" if j==0 else f"={'C' if j==1 else 'D'}{cf_base+4}")
        ws.cell(row=chart_base+5,column=col,value=f"=B8" if j==0 else f"={'C' if j==1 else 'D'}8")

    # Chart 1: Revenue vs Expenses vs Net CF
    chart1=BarChart()
    chart1.type="col"; chart1.title="Revenue vs Expenses vs Net Cash Flow"
    chart1.style=10; chart1.y_axis.title="€"; chart1.x_axis.title="Year"
    chart1.width=16; chart1.height=12
    cats=Reference(ws,min_col=2,max_col=4,min_row=chart_base+1)
    for r_off in [chart_base+2,chart_base+3,chart_base+4]:
        data=Reference(ws,min_col=2,max_col=4,min_row=r_off)
        chart1.add_data(data)
    chart1.set_categories(cats)
    if len(chart1.series)>0: chart1.series[0].title=openpyxl.chart.series.SeriesLabel(v="Revenue")
    if len(chart1.series)>1: chart1.series[1].title=openpyxl.chart.series.SeriesLabel(v="Expenses")
    if len(chart1.series)>2: chart1.series[2].title=openpyxl.chart.series.SeriesLabel(v="Net CF")
    if len(chart1.series)>0: chart1.series[0].graphicalProperties.solidFill=C_MED_BLUE
    if len(chart1.series)>1: chart1.series[1].graphicalProperties.solidFill=C_RED
    ws.add_chart(chart1,f"A{chart_base+7}")

    # Chart 2: AUM Growth
    chart2=AreaChart()
    chart2.title="AUM Growth (€)"; chart2.style=10
    chart2.y_axis.title="€"; chart2.x_axis.title="Year"
    chart2.width=16; chart2.height=12
    cats2=Reference(ws,min_col=2,max_col=4,min_row=chart_base+1)
    data2=Reference(ws,min_col=2,max_col=4,min_row=chart_base+5)
    chart2.add_data(data2); chart2.set_categories(cats2)
    if len(chart2.series)>0:
        chart2.series[0].title=openpyxl.chart.series.SeriesLabel(v="AUM")
        chart2.series[0].graphicalProperties.solidFill=C_GREEN
    ws.add_chart(chart2,f"E{chart_base+7}")

    print("  ✓ Sheet 3: 3-Year Operating Plan")
    return ws, tot_exp_r


# ═══════════════════════════════════════════════════════════════════════
# SHEET 4 – PORTFOLIO CONSTRUCTION
# ═══════════════════════════════════════════════════════════════════════
def build_portfolio(wb):
    ws = wb.create_sheet("Portfolio Construction")
    ws.sheet_properties.tabColor = "FF6F00"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 12

    banner(ws,1,1,"PORTFOLIO CONSTRUCTION",span=11,size=14); ws.row_dimensions[1].height=28
    ws.row_dimensions[2].height=6

    # ── A: Parameters ──
    section(ws,3,1,"A. PORTFOLIO PARAMETERS",span=5); ws.row_dimensions[3].height=20
    th(ws,4,1,"Parameter",span=1); th(ws,4,2,"Value",span=1); th(ws,4,3,"Notes",span=3)
    port_params=[
        ("Total Fund Size (€)","=FundSize",FMT_EUR,False,"Linked from Assumptions"),
        ("Target Companies","='Assumptions & Inputs'!B19",FMT_NUM,True,"Range 15–20"),
        ("Follow-On Reserve (%)","='Assumptions & Inputs'!B23",FMT_PCT,False,"30% of fund"),
        ("Capital for Initial Investments (€)","=B5*(1-B7)",FMT_EUR,False,"Fund × (1 − Follow-On %)"),
        ("Avg Initial Ticket (€)","=B8/B6",FMT_EUR,False,"Initial Capital ÷ # Companies"),
        ("Avg Follow-On per Company (€)","=B5*B7/B6",FMT_EUR,False,"Follow-On Capital ÷ # Companies"),
    ]
    for i,(p,v,fmt,editable,note) in enumerate(port_params):
        r=5+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        lbl(ws,r,1,p); ws.cell(row=r,column=1).fill=fl; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=1)
        c=ws.cell(row=r,column=2,value=v); c.number_format=fmt; c.alignment=align("right","center"); c.border=border()
        if editable: c.fill=inp_fill(); c.font=inp_font()
        else: c.fill=fl; c.font=font(size=10)
        ws.cell(row=r,column=3,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=3).border=border()
        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=5)

    ws.row_dimensions[11].height=6

    # ── B: Portfolio Company Table ──
    section(ws,12,1,"B. PORTFOLIO COMPANIES",span=11); ws.row_dimensions[12].height=20
    headers=["#","Company Name","Sector","Stage","Inv. Date","Initial Inv. (€)","Follow-On (€)","Total Inv. (€)","Curr. Val. (€)","MOIC","Status"]
    for col,hd in enumerate(headers,1):
        th(ws,13,col,hd)

    # Data validation dropdowns
    dv_sector=DataValidation(type="list",formula1='"Longevity,Space"',allow_blank=True)
    dv_stage=DataValidation(type="list",formula1='"Pre-seed,Seed,Series A"',allow_blank=True)
    dv_status=DataValidation(type="list",formula1='"Active,Exited,Failed"',allow_blank=True)
    ws.add_data_validation(dv_sector)
    ws.add_data_validation(dv_stage)
    ws.add_data_validation(dv_status)

    for i in range(20):
        r=14+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        # # column
        c=ws.cell(row=r,column=1,value=i+1); c.fill=fl; c.font=font(size=9); c.alignment=align("center","center"); c.border=border()
        # Company name
        c=ws.cell(row=r,column=2); c.fill=inp_fill(); c.border=border()
        # Sector (dropdown)
        c=ws.cell(row=r,column=3); c.fill=inp_fill(); c.border=border()
        dv_sector.add(c)
        # Stage (dropdown)
        c=ws.cell(row=r,column=4); c.fill=inp_fill(); c.border=border()
        dv_stage.add(c)
        # Date
        c=ws.cell(row=r,column=5); c.fill=inp_fill(); c.border=border(); c.number_format="MMM-YYYY"
        # Initial Investment
        c=ws.cell(row=r,column=6,value=0); c.fill=inp_fill(); c.font=inp_font(); c.number_format=FMT_EUR; c.alignment=align("right","center"); c.border=border()
        # Follow-On
        c=ws.cell(row=r,column=7,value=0); c.fill=inp_fill(); c.font=inp_font(); c.number_format=FMT_EUR; c.alignment=align("right","center"); c.border=border()
        # Total (calculated)
        c=ws.cell(row=r,column=8,value=f"=F{r}+G{r}"); c.fill=calc_fill(); c.number_format=FMT_EUR; c.alignment=align("right","center"); c.border=border()
        # Current Valuation
        c=ws.cell(row=r,column=9,value=0); c.fill=inp_fill(); c.font=inp_font(); c.number_format=FMT_EUR; c.alignment=align("right","center"); c.border=border()
        # MOIC
        c=ws.cell(row=r,column=10,value=f"=IF(H{r}>0,I{r}/H{r},0)"); c.fill=calc_fill(); c.number_format=FMT_MULT; c.alignment=align("right","center"); c.border=border()
        # Status
        c=ws.cell(row=r,column=11); c.fill=inp_fill(); c.border=border()
        dv_status.add(c)

    # Summary rows below table
    sum_base=35; ws.row_dimensions[sum_base].height=6
    section(ws,sum_base+1,1,"C. PORTFOLIO SUMMARY",span=5); ws.row_dimensions[sum_base+1].height=18
    sum_rows=[
        ("Total Companies",         f'=COUNTA(B14:B33)',          FMT_NUM),
        ("Total Deployed (€)",      f"=SUM(H14:H33)",             FMT_EUR),
        ("Longevity Companies",     f'=COUNTIF(C14:C33,"Longevity")',FMT_NUM),
        ("Space Companies",         f'=COUNTIF(C14:C33,"Space")',  FMT_NUM),
        ("Current Portfolio Value (€)",f"=SUM(I14:I33)",           FMT_EUR),
        ("PORTFOLIO MOIC",          f"=IF(I{sum_base+3}>0,I{sum_base+7}/I{sum_base+3},0)",FMT_MULT),
    ]
    # Fix MOIC formula references
    for i,(lbl_txt,formula,fmt) in enumerate(sum_rows):
        r=sum_base+2+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,lbl_txt,bold=(i==5))
        is_key=(i==5)
        c=ws.cell(row=r,column=2,value=formula); c.number_format=fmt
        c.alignment=align("right","center"); c.border=border()
        if is_key:
            c.fill=fill("C8E6C9"); c.font=font(bold=True,size=14,color="1B5E20")
        else:
            c.fill=calc_fill(); c.font=font(size=10)

    # Fix Portfolio MOIC to reference correct cells
    ws.cell(row=sum_base+7,column=2).value=f"=IF(SUM(H14:H33)>0,SUM(I14:I33)/SUM(H14:H33),0)"

    # Named helper cells for Dashboard
    ws.cell(row=sum_base+3,column=2).value="=SUM(H14:H33)"   # Total Deployed → used by Dashboard
    ws.cell(row=sum_base+4,column=2).value=f'=COUNTA(B14:B33)'  # Total Companies

    # ── D: Deployment Schedule ──
    dep_base=sum_base+10; ws.row_dimensions[dep_base].height=6
    section(ws,dep_base+1,1,"D. DEPLOYMENT SCHEDULE",span=11); ws.row_dimensions[dep_base+1].height=18
    for col,hd in enumerate(["Year","Target Capital (€)","# Companies","Avg Ticket (€)","Notes"],1):
        th(ws,dep_base+2,col,hd)
    dep=[
        ("Year 1",  500000,  "3–4", "=B{0}/3.5".format(dep_base+3),  "First investments"),
        ("Year 2", 1500000,  "5–6", "=B{0}/5.5".format(dep_base+4),  "Acceleration"),
        ("Year 3", 2000000,  "5–6", "=B{0}/5.5".format(dep_base+5),  "Main deployment"),
        ("Year 4", 1000000,  "3–4", "=B{0}/3.5".format(dep_base+6),  "Follow-ons & final investments"),
    ]
    for i,(yr,cap,co,avg,note) in enumerate(dep):
        r=dep_base+3+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        ws.cell(row=r,column=1,value=yr).font=font(size=10); ws.cell(row=r,column=1).fill=fl; ws.cell(row=r,column=1).border=border()
        c=ws.cell(row=r,column=2,value=cap); c.number_format=FMT_EUR; c.fill=inp_fill(); c.font=inp_font(); c.alignment=align("right","center"); c.border=border()
        ws.cell(row=r,column=3,value=co).font=font(size=10); ws.cell(row=r,column=3).fill=fl; ws.cell(row=r,column=3).border=border(); ws.cell(row=r,column=3).alignment=align("center","center")
        c=ws.cell(row=r,column=4,value=avg); c.number_format=FMT_EUR; c.fill=calc_fill(); c.font=font(size=10); c.alignment=align("right","center"); c.border=border()
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()
    # Total row
    tot_dep_r=dep_base+7; ws.row_dimensions[tot_dep_r].height=18
    ws.cell(row=tot_dep_r,column=1,value="TOTAL").font=total_font(); ws.cell(row=tot_dep_r,column=1).border=border("medium","medium","medium","medium"); ws.cell(row=tot_dep_r,column=1).fill=fill("D9E1F2")
    c=ws.cell(row=tot_dep_r,column=2,value=f"=SUM(B{dep_base+3}:B{dep_base+6})"); c.number_format=FMT_EUR; c.font=total_font(); c.alignment=align("right","center"); c.border=border("medium","medium","medium","medium"); c.fill=fill("D9E1F2")

    # Charts – Sector allocation pie
    chart_base_p=tot_dep_r+2
    # Hidden chart data for sector/stage
    ws.cell(row=chart_base_p,column=1,value="Longevity").font=font(size=9)
    ws.cell(row=chart_base_p,column=2,value=f'=COUNTIF(C14:C33,"Longevity")').number_format=FMT_NUM
    ws.cell(row=chart_base_p+1,column=1,value="Space").font=font(size=9)
    ws.cell(row=chart_base_p+1,column=2,value=f'=COUNTIF(C14:C33,"Space")').number_format=FMT_NUM

    chart_pie=PieChart()
    chart_pie.title="Sector Allocation"; chart_pie.style=10
    chart_pie.width=14; chart_pie.height=12
    data_p=Reference(ws,min_col=2,max_col=2,min_row=chart_base_p,max_row=chart_base_p+1)
    cats_p=Reference(ws,min_col=1,max_col=1,min_row=chart_base_p,max_row=chart_base_p+1)
    chart_pie.add_data(data_p); chart_pie.set_categories(cats_p)
    ws.add_chart(chart_pie,f"A{chart_base_p+3}")

    print("  ✓ Sheet 4: Portfolio Construction")
    return ws


# ═══════════════════════════════════════════════════════════════════════
# SHEET 5 – SCENARIO ANALYSIS
# ═══════════════════════════════════════════════════════════════════════
def build_scenarios(wb):
    ws = wb.create_sheet("Scenario Analysis")
    ws.sheet_properties.tabColor = C_RED
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30

    banner(ws,1,1,"SCENARIO ANALYSIS",span=5,size=14); ws.row_dimensions[1].height=28
    ws.row_dimensions[2].height=6

    # ── A: Scenario Inputs ──
    section(ws,3,1,"A. SCENARIO INPUTS",span=5); ws.row_dimensions[3].height=20
    for col,hd in enumerate(["Parameter","Best Case","Base Case","Worst Case","Notes"],1):
        th(ws,4,col,hd)

    scen_inp=[
        ("Fund Size (€)",             5000000, 3000000, 1000000, FMT_EUR, "Total AUM raised"),
        ("Portfolio MOIC",            5.0,     3.0,     1.0,     FMT_MULT,"Gross fund multiple"),
        ("Years to Exit (avg)",       6,       7,       9,       FMT_NUM, "Average holding period"),
        ("Avg Annual Op Expenses (€)",60000,   70000,   80000,   FMT_EUR, "Total annual expenses"),
        ("Management Fee (%)",        0.015,   0.015,   0.015,   FMT_PCT, "Fixed at 1.5%"),
        ("Carried Interest (%)",      0.25,    0.25,    0.25,    FMT_PCT, "Fixed at 25%"),
    ]
    for i,(p,best,base,worst,fmt,note) in enumerate(scen_inp):
        r=5+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        lbl(ws,r,1,p); ws.cell(row=r,column=1).fill=fl
        for j,v in enumerate([best,base,worst],2):
            c=ws.cell(row=r,column=j,value=v); c.number_format=fmt
            c.alignment=align("right","center"); c.border=border()
            c.fill=inp_fill(); c.font=inp_font()
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()
    # B5=BestFundSize, C5=BaseFundSize, D5=WorstFundSize
    # B6=BestMOIC, C6=BaseMOIC, D6=WorstMOIC
    # B7=BestYears, C7=BaseYears, D7=WorstYears

    ws.row_dimensions[11].height=6

    # ── B: Scenario Outcomes ──
    section(ws,12,1,"B. SCENARIO OUTCOMES",span=5); ws.row_dimensions[12].height=20
    for col,hd in enumerate(["Metric","Best Case","Base Case","Worst Case","Formula basis"],1):
        th(ws,13,col,hd)

    outcomes=[
        ("Fund Returns (Gross)","=B5*B6","=C5*C6","=D5*D6",FMT_EUR,"Fund Size × MOIC"),
        ("Gross Profit","=B14-B5","=C14-C5","=D14-D5",FMT_EUR,"Returns − Fund Size"),
        ("Hurdle to LPs (8% comp.)","=B5*((1+B10)^B7)-B5","=C5*((1+C10)^C7)-C5","=D5*((1+D10)^D7)-D5",FMT_EUR,"Fund × (1.08^yrs) − Fund"),
        ("Profit Above Hurdle","=MAX(0,B15-B16)","=MAX(0,C15-C16)","=MAX(0,D15-D16)",FMT_EUR,"MAX(0, Profit − Hurdle)"),
        ("GP Carry (25%)","=B17*B11","=C17*C11","=D17*D11",FMT_EUR,"Excess × 25%"),
        ("LP Net Returns","=B5+B16+(B17*0.75)","=C5+C16+(C17*0.75)","=D5+D16+(D17*0.75)",FMT_EUR,"Capital+Hurdle+75%Excess"),
        ("LP Net Multiple","=B19/B5","=C19/C5","=D19/D5",FMT_MULT,"LP Returns ÷ Fund Size"),
        ("LP Net IRR (approx.)","=(B19/B5)^(1/B7)-1","=(C19/C5)^(1/C7)-1","=(D19/D5)^(1/D7)-1",FMT_PCT,"(Multiple)^(1/Years)−1"),
        ("Mgmt Fees (10 years)","=B5*B10*10","=C5*C10*10","=D5*D10*10",FMT_EUR,"Fund × 1.5% × 10"),
        ("GP Total Compensation","=B22+B18","=C22+C18","=D22+D18",FMT_EUR,"Mgmt Fees + Carry"),
    ]
    for i,(metric,best,base,worst,fmt,note) in enumerate(outcomes):
        r=14+i; ws.row_dimensions[r].height=18
        is_key=(i in [6,7,9])
        lbl(ws,r,1,metric,bold=is_key)
        for j,v in enumerate([best,base,worst],2):
            c=ws.cell(row=r,column=j,value=v); c.number_format=fmt
            c.alignment=align("right","center"); c.border=border()
            if is_key:
                c.fill=fill("C8E6C9"); c.font=font(bold=True,size=11,color="1B5E20")
            else:
                c.fill=calc_fill(); c.font=font(size=10)
        ws.cell(row=r,column=5,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=5).border=border()

    # Conditional formatting on IRR row (row 21)
    irr_row=21
    cf_range=f"B{irr_row}:D{irr_row}"
    ws.conditional_formatting.add(cf_range,CellIsRule(operator="greaterThan",formula=["0.20"],fill=fill("C8E6C9"),font=font(color="1B5E20")))
    ws.conditional_formatting.add(cf_range,CellIsRule(operator="between",formula=["0.10","0.20"],fill=fill("FFF9C4"),font=font(color="F57F17")))
    ws.conditional_formatting.add(cf_range,CellIsRule(operator="lessThan",formula=["0.10"],fill=fill(C_LT_RED),font=font(color=C_RED)))

    ws.row_dimensions[24].height=6

    # ── C: Comparison Table ──
    section(ws,25,1,"C. SCENARIO COMPARISON SUMMARY",span=5); ws.row_dimensions[25].height=20
    for col,hd in enumerate(["Metric","Best Case","Base Case","Worst Case"],1):
        th(ws,26,col,hd)
    comp=[
        ("Fund Size (€)",       "=B5","=C5","=D5",FMT_EUR),
        ("Gross Returns (€)",   "=B14","=C14","=D14",FMT_EUR),
        ("LP Net Multiple",     "=B20","=C20","=D20",FMT_MULT),
        ("LP Net IRR",          "=B21","=C21","=D21",FMT_PCT),
        ("GP Total Comp (€)",   "=B23","=C23","=D23",FMT_EUR),
        ("Mgmt Fees 10yr (€)",  "=B22","=C22","=D22",FMT_EUR),
        ("GP Carry (€)",        "=B18","=C18","=D18",FMT_EUR),
    ]
    for i,(metric,best,base,worst,fmt) in enumerate(comp):
        r=27+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        lbl(ws,r,1,metric); ws.cell(row=r,column=1).fill=fl
        for j,v in enumerate([best,base,worst],2):
            c=ws.cell(row=r,column=j,value=v); c.number_format=fmt; c.fill=fl
            c.alignment=align("right","center"); c.border=border(); c.font=font(size=10)

    ws.row_dimensions[34].height=6

    # ── D: Sensitivity Table (static approximation) ──
    section(ws,35,1,"D. SENSITIVITY — LP NET IRR BY FUND SIZE & MOIC",span=5); ws.row_dimensions[35].height=20
    ws.cell(row=36,column=1,value="Rows = Portfolio MOIC | Columns = Fund Size (for reference, IRR is size-independent)").font=font(italic=True,size=9,color=C_GRAY)
    ws.row_dimensions[36].height=14
    # Column headers
    th(ws,37,1,"MOIC \\ Years")
    for j,yrs in enumerate([5,6,7,8,9,10],2):
        th(ws,37,j,f"{yrs} yrs")
    moic_vals=[1.0,1.5,2.0,2.5,3.0,3.5,4.0,5.0,7.0,10.0]
    for i,moic in enumerate(moic_vals):
        r=38+i; ws.row_dimensions[r].height=16; fl=alt_fill(i)
        ws.cell(row=r,column=1,value=f"{moic}x").font=font(bold=True,size=10); ws.cell(row=r,column=1).fill=fl; ws.cell(row=r,column=1).border=border()
        for j,yrs in enumerate([5,6,7,8,9,10],2):
            irr_val = moic**(1/yrs)-1
            c=ws.cell(row=r,column=j,value=irr_val); c.number_format=FMT_PCT; c.alignment=align("right","center"); c.border=border()
            # Colour-code
            if irr_val>=0.20: c.fill=fill("C8E6C9"); c.font=font(color="1B5E20",size=10)
            elif irr_val>=0.10: c.fill=fill("FFF9C4"); c.font=font(color="F57F17",size=10)
            else: c.fill=fill(C_LT_RED); c.font=font(color=C_RED,size=10)

    ws.row_dimensions[48].height=6

    # ── Charts ──
    chart_base_s=49
    # Hidden data for scenario charts
    for col,hd in enumerate(["","Best Case","Base Case","Worst Case"],1):
        ws.cell(row=chart_base_s,column=col,value=hd).font=font(bold=True,size=9)
    ws.cell(row=chart_base_s+1,column=1,value="LP Net Multiple").font=font(size=9)
    ws.cell(row=chart_base_s+2,column=1,value="LP Net IRR").font=font(size=9)
    ws.cell(row=chart_base_s+3,column=1,value="GP Mgmt Fees").font=font(size=9)
    ws.cell(row=chart_base_s+4,column=1,value="GP Carry").font=font(size=9)
    for j in range(3):
        col=j+2
        mult_ref=["=B20","=C20","=D20"][j]
        irr_ref=["=B21","=C21","=D21"][j]
        mgmt_ref=["=B22","=C22","=D22"][j]
        carry_ref=["=B18","=C18","=D18"][j]
        ws.cell(row=chart_base_s+1,column=col,value=mult_ref)
        ws.cell(row=chart_base_s+2,column=col,value=irr_ref)
        ws.cell(row=chart_base_s+3,column=col,value=mgmt_ref)
        ws.cell(row=chart_base_s+4,column=col,value=carry_ref)

    # Chart: LP Returns by Scenario
    chart_scen=BarChart()
    chart_scen.type="col"; chart_scen.title="LP Net Multiple by Scenario"
    chart_scen.style=10; chart_scen.y_axis.title="Net Multiple (x)"
    chart_scen.width=14; chart_scen.height=12
    cats_s=Reference(ws,min_col=2,max_col=4,min_row=chart_base_s)
    data_s=Reference(ws,min_col=2,max_col=4,min_row=chart_base_s+1)
    chart_scen.add_data(data_s); chart_scen.set_categories(cats_s)
    if len(chart_scen.series)>0: chart_scen.series[0].graphicalProperties.solidFill=C_MED_BLUE
    ws.add_chart(chart_scen,f"A{chart_base_s+6}")

    # Chart: GP Economics by Scenario (stacked)
    chart_gp=BarChart()
    chart_gp.type="col"; chart_gp.title="GP Economics by Scenario (€)"
    chart_gp.style=10; chart_gp.y_axis.title="€"
    chart_gp.grouping="stacked"; chart_gp.width=14; chart_gp.height=12
    data_mgmt=Reference(ws,min_col=2,max_col=4,min_row=chart_base_s+3)
    data_carry=Reference(ws,min_col=2,max_col=4,min_row=chart_base_s+4)
    chart_gp.add_data(data_mgmt); chart_gp.add_data(data_carry)
    chart_gp.set_categories(cats_s)
    if len(chart_gp.series)>0: chart_gp.series[0].title=openpyxl.chart.series.SeriesLabel(v="Mgmt Fees"); chart_gp.series[0].graphicalProperties.solidFill=C_MED_BLUE
    if len(chart_gp.series)>1: chart_gp.series[1].title=openpyxl.chart.series.SeriesLabel(v="Carry"); chart_gp.series[1].graphicalProperties.solidFill=C_GREEN
    ws.add_chart(chart_gp,f"E{chart_base_s+6}")

    print("  ✓ Sheet 5: Scenario Analysis")
    return ws


# ═══════════════════════════════════════════════════════════════════════
# SHEET 6 – INVESTOR CALCULATOR
# ═══════════════════════════════════════════════════════════════════════
def build_investor_calc(wb):
    ws = wb.create_sheet("Investor Calculator")
    ws.sheet_properties.tabColor = C_GREEN
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20

    banner(ws,1,1,"INVESTOR RETURNS CALCULATOR",span=6,size=14); ws.row_dimensions[1].height=28
    ws.cell(row=2,column=1,value="Interactive tool — change the yellow cells to model your potential returns").font=font(italic=True,size=10,color=C_GRAY)
    ws.row_dimensions[2].height=16; ws.row_dimensions[3].height=6

    # ── A: Your Investment ──
    section(ws,4,1,"A. YOUR INVESTMENT (Edit yellow cells)",span=6); ws.row_dimensions[4].height=20
    th(ws,5,1,"Input"); th(ws,5,2,"Your Value"); th(ws,5,4,"Guidance")

    inv_inputs=[
        ("Investment Amount (€)",    10000, FMT_EUR,  "Minimum €10,000"),
        ("Expected Portfolio MOIC",  3.0,   FMT_MULT, "Dropdown: 1x–10x"),
        ("Holding Period (years)",   7,     FMT_NUM,  "Range: 5–10 years"),
    ]
    for i,(p,v,fmt,note) in enumerate(inv_inputs):
        r=6+i; ws.row_dimensions[r].height=20
        lbl(ws,r,1,p)
        inp(ws,r,2,v,fmt=fmt)
        ws.cell(row=r,column=4,value=note).font=font(size=9,italic=True,color=C_GRAY); ws.cell(row=r,column=4).border=border()
        ws.merge_cells(start_row=r,start_column=4,end_row=r,end_column=6)
    # B6=Investment, B7=MOIC, B8=Years

    # Data validations
    dv_moic=DataValidation(type="list",formula1='"1x,1.5x,2x,2.5x,3x,3.5x,4x,4.5x,5x,6x,7x,8x,9x,10x"',allow_blank=False)
    ws.add_data_validation(dv_moic)
    dv_inv=DataValidation(type="whole",operator="greaterThanOrEqual",formula1="10000",showErrorMessage=True,error="Minimum investment is €10,000")
    ws.add_data_validation(dv_inv); dv_inv.add(ws["B6"])
    dv_yrs=DataValidation(type="whole",operator="between",formula1="5",formula2="10",showErrorMessage=True,error="Enter years between 5 and 10")
    ws.add_data_validation(dv_yrs); dv_yrs.add(ws["B8"])

    ws.row_dimensions[9].height=8

    # ── B: Your Returns ──
    section(ws,10,1,"B. YOUR PROJECTED RETURNS",span=6); ws.row_dimensions[10].height=20
    ws.cell(row=11,column=1,value="All calculations use the 8% hurdle rate and 25% carried interest structure of the fund").font=font(italic=True,size=9,color=C_GRAY)
    ws.row_dimensions[11].height=14; ws.row_dimensions[12].height=6

    ret_calcs=[
        ("Initial Investment",               "=B6",                               FMT_EUR,  False),
        ("Gross Return (before fees)",        "=B6*B7",                            FMT_EUR,  False),
        ("Gross Profit",                      "=B14-B6",                           FMT_EUR,  False),
        ("","","",""),
        ("Preferred Return (8% hurdle)",      "=B6*((1.08^B8)-1)",                 FMT_EUR,  False),
        ("You Receive First (Capital+Hurdle)","=B6+B17",                           FMT_EUR,  False),
        ("","","",""),
        ("Remaining Profit Above Hurdle",     "=MAX(0,B15-B17)",                   FMT_EUR,  False),
        ("Your 75% Share",                    "=B20*0.75",                         FMT_EUR,  False),
        ("GP Carry 25%",                      "=B20*0.25",                         FMT_EUR,  False),
        ("","","",""),
        ("YOUR TOTAL DISTRIBUTION",           "=B18+B21",                          FMT_EUR,  True),
        ("YOUR NET MULTIPLE",                 "=B23/B6",                           FMT_MULT, True),
        ("YOUR NET IRR (approx.)",            "=(B23/B6)^(1/B8)-1",               FMT_PCT,  True),
    ]
    for i,(lbl_txt,formula,fmt,is_key) in enumerate(ret_calcs):
        r=13+i; ws.row_dimensions[r].height=18 if lbl_txt else 8
        if lbl_txt:
            lbl(ws,r,1,lbl_txt,bold=is_key)
            c=ws.cell(row=r,column=2,value=formula); c.number_format=fmt
            c.alignment=align("right","center"); c.border=border()
            if is_key:
                c.fill=fill("C8E6C9"); c.font=font(bold=True,size=16,color="1B5E20")
                ws.row_dimensions[r].height=28
            else:
                c.fill=calc_fill(); c.font=font(size=10)

    ws.row_dimensions[27].height=8

    # ── C: Comparison Table ──
    section(ws,28,1,"C. SCENARIO COMPARISON (based on your investment amount)",span=6); ws.row_dimensions[28].height=20
    for col,hd in enumerate(["Scenario","Portfolio MOIC","Your Investment","Your Distribution","Your Multiple","Your IRR"],1):
        th(ws,29,col,hd)

    comp_scen=[
        ("Conservative","=1*1",  2.0),
        ("Base Case",   "=1*1",  3.0),
        ("Optimistic",  "=1*1",  5.0),
    ]
    for i,(name,_,moic) in enumerate(comp_scen):
        r=30+i; ws.row_dimensions[r].height=18; fl=alt_fill(i)
        ws.cell(row=r,column=1,value=name).font=font(size=10); ws.cell(row=r,column=1).fill=fl; ws.cell(row=r,column=1).border=border()
        ws.cell(row=r,column=2,value=moic).number_format=FMT_MULT; ws.cell(row=r,column=2).fill=fl; ws.cell(row=r,column=2).font=font(size=10); ws.cell(row=r,column=2).alignment=align("right","center"); ws.cell(row=r,column=2).border=border()
        inv_ref="$B$6"; yrs_ref="$B$8"
        gross=f"{inv_ref}*B{r}"
        hurdle=f"{inv_ref}*((1.08^{yrs_ref})-1)"
        pah=f"MAX(0,{gross}-{inv_ref}-{hurdle})"
        dist=f"={inv_ref}+{hurdle}+({pah}*0.75)"
        for j,v in enumerate([f"={inv_ref}",dist,f"=E{r}/{inv_ref}",f"=(E{r}/{inv_ref})^(1/{yrs_ref})-1"],3):
            c=ws.cell(row=r,column=j,value=v); c.fill=fl; c.font=font(size=10); c.alignment=align("right","center"); c.border=border()
        ws.cell(row=r,column=3).number_format=FMT_EUR
        ws.cell(row=r,column=4).number_format=FMT_EUR
        ws.cell(row=r,column=5).number_format=FMT_MULT
        ws.cell(row=r,column=6).number_format=FMT_PCT

    ws.row_dimensions[33].height=8

    # ── D: Fee Impact Analysis ──
    section(ws,34,1,"D. FEE IMPACT ANALYSIS",span=6); ws.row_dimensions[34].height=20
    th(ws,35,1,"Item"); th(ws,35,2,"Amount")
    fee_items=[
        ("Gross Returns (before fees)",     "=B6*B7",              FMT_EUR,  False),
        ("Gross Multiple",                  "=B7",                 FMT_MULT, False),
        ("Mgmt Fees Impact (1.5% × 10yr)","=B6*0.015*10",         FMT_EUR,  False),
        ("Carry Paid to GP",               "=B22",                FMT_EUR,  False),
        ("Net to You",                     "=B23",                FMT_EUR,  True),
        ("Net Multiple",                   "=B24",                FMT_MULT, True),
        ("Blended Fee Rate",               "=IF(B36>0,(B38+B39)/B36,0)",FMT_PCT,False),
    ]
    for i,(lbl_txt,formula,fmt,is_key) in enumerate(fee_items):
        r=36+i; ws.row_dimensions[r].height=18
        lbl(ws,r,1,lbl_txt,bold=is_key)
        c=ws.cell(row=r,column=2,value=formula); c.number_format=fmt
        c.alignment=align("right","center"); c.border=border()
        c.fill=fill("C8E6C9") if is_key else calc_fill()
        c.font=font(bold=is_key,size=11 if is_key else 10,color="1B5E20" if is_key else "000000")

    ws.row_dimensions[43].height=8

    # ── F: Risk Disclaimer ──
    section(ws,44,1,"IMPORTANT RISK DISCLAIMER",span=6,clr=C_RED); ws.row_dimensions[44].height=20
    disclaimer=("This calculator provides illustrative projections only. Actual returns may vary significantly. "
                "Venture capital investments are high-risk and illiquid. You could lose your entire investment. "
                "Past performance (if any) does not guarantee future results. "
                "MOIC and IRR targets are aspirational, not promises. "
                "Inflection Ventures Fund I is not yet regulated or approved by any authority.")
    ws.merge_cells("A45:F48")
    c=ws["A45"]; c.value=disclaimer; c.font=font(size=10,italic=True,color="7B0000")
    c.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    c.fill=fill("FFEBEE"); c.border=border("medium","medium","medium","medium")
    for r in range(45,49): ws.row_dimensions[r].height=18

    print("  ✓ Sheet 6: Investor Calculator")
    return ws


# ═══════════════════════════════════════════════════════════════════════
# SHEET 1 – DASHBOARD (built last so all formulas exist)
# ═══════════════════════════════════════════════════════════════════════
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = C_DARK_BLUE
    ws.sheet_view.showGridLines = False

    # Column widths
    col_w=[2,22,18,22,18,22,18,22,18,2]
    for i,w in enumerate(col_w,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    # ── Title Banner ──
    ws.row_dimensions[1].height=8
    ws.row_dimensions[2].height=55
    ws.merge_cells("B2:I2")
    c=ws["B2"]
    c.value="INFLECTION VENTURES FUND I\nFinancial Model  ·  Executive Summary"
    c.fill=fill(C_DARK_BLUE); c.font=Font(name="Calibri",bold=True,size=22,color=C_WHITE)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

    ws.row_dimensions[3].height=22
    ws.merge_cells("B3:F3")
    c=ws["B3"]; c.value="Luxembourg ELTIF 2.0  ·  Longevity & Space Economy  ·  Launch: Q2 2026"
    c.fill=fill(C_MED_BLUE); c.font=font(size=11,color=C_WHITE); c.alignment=align("left","center")

    ws.merge_cells("G3:H3")
    c=ws["G3"]; c.value="Status:"
    c.fill=fill(C_MED_BLUE); c.font=font(bold=True,size=11,color=C_WHITE); c.alignment=align("right","center")
    c=ws["I3"]; c.value="Fundraising"
    c.fill=inp_fill(); c.font=font(bold=True,size=11); c.alignment=align("center","center"); c.border=border()
    dv=DataValidation(type="list",formula1='"Fundraising,Deployed,Exited"'); ws.add_data_validation(dv); dv.add(ws["I3"])

    ws.row_dimensions[4].height=8

    # ── 4 KPI Boxes (2 rows × 2 cols = actually 4 boxes, side by side) ──
    # Box layout: cols B-C, D-E, F-G, H-I  rows 5-11
    box_titles=["CAPITAL","INVESTORS","PORTFOLIO","PERFORMANCE (Projected)"]
    box_cols=[(2,3),(4,5),(6,7),(8,9)]

    for idx,((c1,c2),title) in enumerate(zip(box_cols,box_titles)):
        ws.row_dimensions[5].height=22
        ws.merge_cells(start_row=5,start_column=c1,end_row=5,end_column=c2)
        h=ws.cell(row=5,column=c1,value=title)
        h.fill=fill(C_DARK_BLUE); h.font=font(bold=True,size=12,color=C_WHITE)
        h.alignment=align("center","center"); h.border=border("medium","thin","medium","medium")

    # Capital box (B-C, rows 6-10)
    capital=[
        ("Target Fund Size","€3.0M – €5.0M","@"),
        ("Capital Raised","='3-Year Operating Plan'!B6+'3-Year Operating Plan'!C6+'3-Year Operating Plan'!D6",FMT_EUR),
        ("Capital Deployed","=ABS('3-Year Operating Plan'!B7+'3-Year Operating Plan'!C7+'3-Year Operating Plan'!D7)",FMT_EUR),
        ("Available Capital","=B8-B9","@"),
        ("Min Investment","€10,000","@"),
    ]
    inv_box=[
        ("Target Investors","300 – 500","@"),
        ("Avg Commitment","=FundSize/350",FMT_EUR),
        ("Min Investment","€10,000","@"),
        ("Fund Duration","10 years","@"),
        ("Structure","ELTIF 2.0","@"),
    ]
    port_box=[
        ("Target Companies","15 – 20","@"),
        ("Current Companies","='Portfolio Construction'!B37",FMT_NUM),
        ("Total Deployed","='Portfolio Construction'!B36",FMT_EUR),
        ("Portfolio MOIC","='Portfolio Construction'!B40",FMT_MULT),
        ("Avg Initial Ticket","='Assumptions & Inputs'!B22",FMT_EUR),
    ]
    perf_box=[
        ("Target MOIC","3x – 5x","@"),
        ("Target IRR","25% – 35%","@"),
        ("Management Fee","1.5%","@"),
        ("Carried Interest","25%","@"),
        ("Hurdle Rate","8.0%","@"),
    ]
    all_boxes=[capital,inv_box,port_box,perf_box]

    for box_idx,(cols,box_data) in enumerate(zip(box_cols,all_boxes)):
        c1,c2=cols
        for row_i,(lbl_txt,val,fmt) in enumerate(box_data):
            r=6+row_i; ws.row_dimensions[r].height=20
            lc=ws.cell(row=r,column=c1,value=lbl_txt)
            lc.fill=fill("EBF1FF"); lc.font=font(size=10); lc.alignment=align("left","center"); lc.border=border("thin","thin","medium","thin")
            vc=ws.cell(row=r,column=c2,value=val)
            vc.fill=calc_fill() if val and str(val).startswith("=") else fill(C_LT_GREEN)
            vc.font=font(bold=True,size=10,color=C_DARK_BLUE); vc.alignment=align("right","center")
            vc.border=border("thin","thin","thin","medium"); vc.number_format=fmt

    # Fix first items that should not show formula errors
    ws["C6"].value="€3.0M – €5.0M"; ws["C6"].fill=fill(C_LT_GREEN)
    ws["C10"].value="€10,000"; ws["C10"].fill=fill(C_LT_GREEN)
    ws["E6"].value="300 – 500"; ws["E6"].fill=fill(C_LT_GREEN)
    ws["E8"].value="€10,000"; ws["E8"].fill=fill(C_LT_GREEN)
    ws["E9"].value="10 years"; ws["E9"].fill=fill(C_LT_GREEN)
    ws["E10"].value="ELTIF 2.0"; ws["E10"].fill=fill(C_LT_GREEN)
    ws["G6"].value="15 – 20"; ws["G6"].fill=fill(C_LT_GREEN)
    ws["I6"].value="3x – 5x"; ws["I6"].fill=fill(C_LT_GREEN)
    ws["I7"].value="25% – 35%"; ws["I7"].fill=fill(C_LT_GREEN)
    ws["I8"].value="1.5%"; ws["I8"].fill=fill(C_LT_GREEN)
    ws["I9"].value="25%"; ws["I9"].fill=fill(C_LT_GREEN)
    ws["I10"].value="8.0%"; ws["I10"].fill=fill(C_LT_GREEN)

    ws.row_dimensions[11].height=8

    # ── Summary Table ──
    ws.row_dimensions[12].height=22
    ws.merge_cells("B12:I12")
    h=ws["B12"]; h.value="3-YEAR FINANCIAL SUMMARY"
    h.fill=fill(C_DARK_BLUE); h.font=font(bold=True,size=13,color=C_WHITE); h.alignment=align("center","center")

    ws.row_dimensions[13].height=20
    for col,hd in enumerate(["Metric","Year 1","Year 2","Year 3"],0):
        c1_s,c2_s=[(2,3),(4,5),(6,7),(8,9)][col]
        ws.merge_cells(start_row=13,start_column=c1_s,end_row=13,end_column=c2_s)
        c=ws.cell(row=13,column=c1_s,value=hd)
        c.fill=fill(C_MED_BLUE); c.font=font(bold=True,size=10,color=C_WHITE); c.alignment=align("center","center"); c.border=border()

    summary_rows=[
        ("AUM (Average, €)",         "='3-Year Operating Plan'!B9","='3-Year Operating Plan'!C9","='3-Year Operating Plan'!D9"),
        ("Management Fees (€)",      "='3-Year Operating Plan'!B12","='3-Year Operating Plan'!C12","='3-Year Operating Plan'!D12"),
        ("Operating Expenses (€)",   "='3-Year Operating Plan'!B27","='3-Year Operating Plan'!C27","='3-Year Operating Plan'!D27"),
        ("Net Cash Flow (€)",        "='3-Year Operating Plan'!B29","='3-Year Operating Plan'!C29","='3-Year Operating Plan'!D29"),
        ("Cumulative Cash Flow (€)", "='3-Year Operating Plan'!B30","='3-Year Operating Plan'!C30","='3-Year Operating Plan'!D30"),
    ]
    for i,(metric,y1,y2,y3) in enumerate(summary_rows):
        r=14+i; ws.row_dimensions[r].height=20; fl=alt_fill(i)
        c1_s,c2_s=2,3
        ws.merge_cells(start_row=r,start_column=c1_s,end_row=r,end_column=c2_s)
        c=ws.cell(row=r,column=c1_s,value=metric)
        c.fill=fl; c.font=font(bold=(i==4),size=10); c.alignment=align("left","center"); c.border=border()
        for j,(ys_cols,val) in enumerate(zip([(4,5),(6,7),(8,9)],[y1,y2,y3])):
            c1_y,c2_y=ys_cols
            ws.merge_cells(start_row=r,start_column=c1_y,end_row=r,end_column=c2_y)
            c=ws.cell(row=r,column=c1_y,value=val)
            c.fill=fl; c.font=font(bold=(i==4),size=10); c.alignment=align("right","center")
            c.border=border(); c.number_format=FMT_EUR

    ws.row_dimensions[19].height=8

    # ── Charts ──
    ws.row_dimensions[20].height=22
    ws.merge_cells("B20:I20")
    h=ws["B20"]; h.value="VISUAL OVERVIEW"
    h.fill=fill(C_DARK_BLUE); h.font=font(bold=True,size=13,color=C_WHITE); h.alignment=align("center","center")

    # Hidden chart data rows 40+
    ws.row_dimensions[40].height=14
    for col,hd in enumerate(["","Year 1","Year 2","Year 3","Year 4"],1):
        ws.cell(row=40,column=col,value=hd).font=font(bold=True,size=9)
    ws.cell(row=41,column=1,value="Deployed (€M)").font=font(size=9)
    for col,v in enumerate([0.5,1.5,2.0,1.0],2):
        ws.cell(row=41,column=col,value=v)

    ws.row_dimensions[43].height=14
    for col,hd in enumerate(["","Longevity","Space"],1):
        ws.cell(row=43,column=col,value=hd).font=font(bold=True,size=9)
    ws.cell(row=44,column=1,value="Allocation").font=font(size=9)
    ws.cell(row=44,column=2,value=0.5); ws.cell(row=44,column=3,value=0.5)

    ws.row_dimensions[46].height=14
    for col,hd in enumerate(["","Year 1","Year 2","Year 3"],1):
        ws.cell(row=46,column=col,value=hd).font=font(bold=True,size=9)
    ws.cell(row=47,column=1,value="Revenue").font=font(size=9)
    ws.cell(row=47,column=2,value="='3-Year Operating Plan'!B12")
    ws.cell(row=47,column=3,value="='3-Year Operating Plan'!C12")
    ws.cell(row=47,column=4,value="='3-Year Operating Plan'!D12")
    ws.cell(row=48,column=1,value="Expenses").font=font(size=9)
    ws.cell(row=48,column=2,value="='3-Year Operating Plan'!B27")
    ws.cell(row=48,column=3,value="='3-Year Operating Plan'!C27")
    ws.cell(row=48,column=4,value="='3-Year Operating Plan'!D27")
    ws.cell(row=49,column=1,value="Net CF").font=font(size=9)
    ws.cell(row=49,column=2,value="='3-Year Operating Plan'!B29")
    ws.cell(row=49,column=3,value="='3-Year Operating Plan'!C29")
    ws.cell(row=49,column=4,value="='3-Year Operating Plan'!D29")

    # Chart 1: Capital Deployment
    ch1=BarChart(); ch1.type="col"; ch1.title="Capital Deployment (€M)"
    ch1.style=10; ch1.y_axis.title="€M"; ch1.width=12; ch1.height=12
    cats1=Reference(ws,min_col=2,max_col=5,min_row=40)
    data1=Reference(ws,min_col=2,max_col=5,min_row=41)
    ch1.add_data(data1); ch1.set_categories(cats1)
    if len(ch1.series)>0: ch1.series[0].graphicalProperties.solidFill=C_DARK_BLUE
    ws.add_chart(ch1,"B21")

    # Chart 2: Sector Allocation
    ch2=PieChart(); ch2.title="Sector Allocation"; ch2.style=10
    ch2.width=11; ch2.height=12
    data2=Reference(ws,min_col=2,max_col=3,min_row=44)
    cats2=Reference(ws,min_col=2,max_col=3,min_row=43)
    ch2.add_data(data2); ch2.set_categories(cats2)
    ws.add_chart(ch2,"E21")

    # Chart 3: Cash Flow
    ch3=LineChart(); ch3.title="Operating Cash Flow"; ch3.style=10
    ch3.y_axis.title="€"; ch3.width=12; ch3.height=12
    cats3=Reference(ws,min_col=2,max_col=4,min_row=46)
    for r_off,lbl_txt in [(47,"Revenue"),(48,"Expenses"),(49,"Net CF")]:
        d=Reference(ws,min_col=2,max_col=4,min_row=r_off)
        ch3.add_data(d)
    ch3.set_categories(cats3)
    if len(ch3.series)>0: ch3.series[0].title=openpyxl.chart.series.SeriesLabel(v="Revenue"); ch3.series[0].graphicalProperties.solidFill=C_MED_BLUE
    if len(ch3.series)>1: ch3.series[1].title=openpyxl.chart.series.SeriesLabel(v="Expenses"); ch3.series[1].graphicalProperties.solidFill=C_RED
    if len(ch3.series)>2: ch3.series[2].title=openpyxl.chart.series.SeriesLabel(v="Net CF"); ch3.series[2].graphicalProperties.solidFill=C_GREEN
    ws.add_chart(ch3,"H21")

    # Version / navigation
    ws.row_dimensions[38].height=18
    ws.merge_cells("B38:D38")
    c=ws["B38"]; c.value="Model Version: 1.0  ·  Last Updated: March 2026"
    c.font=font(size=9,italic=True,color=C_GRAY); c.alignment=align("left","center")

    nav=[("Fund Economics","E38"),("Operating Plan","F38"),("Portfolio","G38"),("Scenarios","H38"),("Inv. Calc","I38")]
    for label,ref in nav:
        c=ws[ref]; c.value=f"→ {label}"
        c.font=Font(name="Calibri",size=9,color=C_DARK_BLUE,underline="single"); c.alignment=align("center","center")

    print("  ✓ Sheet 1: Dashboard")
    return ws

# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════
import openpyxl

def main():
    print("\n🏗  Building Inflection Ventures Financial Model...\n")
    wb = Workbook()
    wb.remove(wb.active)

    # Build in dependency order
    ws_assum, exp_tot_r = build_assumptions(wb)
    build_fund_economics(wb)
    build_operating_plan(wb)
    build_portfolio(wb)
    build_scenarios(wb)
    build_investor_calc(wb)
    build_dashboard(wb)

    # Reorder: Dashboard first (directly reorder internal sheet list)
    sheet_order=["Dashboard","Fund Economics","3-Year Operating Plan",
                 "Portfolio Construction","Scenario Analysis","Investor Calculator","Assumptions & Inputs"]
    sheets_map={s.title:s for s in wb._sheets}
    wb._sheets=[sheets_map[n] for n in sheet_order]

    fname="/home/user/truffle-flattener/Inflection_Ventures_Financial_Model_v1.0.xlsx"
    wb.save(fname)
    print(f"\n✅  Saved: {fname}")
    print(f"\nSheets: {wb.sheetnames}")

if __name__ == "__main__":
    main()
